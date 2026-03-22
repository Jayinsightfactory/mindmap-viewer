#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Orbit AI - 자동화 CLI 도구 (orbit-auto.py)

직원 PC에 배포하여 자동화 스크립트를 테스트/실행하는 CLI 도구.

[필수 패키지]
  pip install requests pyautogui openpyxl pyperclip

[사용법]
  orbit-auto.py --target=nenova --action=order --test     # 드라이런 (입력 시뮬레이션)
  orbit-auto.py --target=nenova --action=order --run      # 실제 실행
  orbit-auto.py --target=excel --action=deduction --file=차감.xlsx
  orbit-auto.py --target=excel --action=shipping --file=출고.xlsx
  orbit-auto.py --status                                   # 서버 상태 확인
  orbit-auto.py --sync                                     # 마스터 DB 동기화
  orbit-auto.py --parse "MEL ROSE CHINA / Catherine : 30"  # 파서 테스트
  orbit-auto.py --watch                                    # 클립보드 감시 모드

Python 3.8+ 필요
"""

import argparse
import json
import logging
import os
import re
import sys
import time
from datetime import datetime
from pathlib import Path

# ---------------------------------------------------------------------------
# 필수 의존성: requests
# ---------------------------------------------------------------------------
try:
    import requests
except ImportError:
    print("[오류] requests 패키지가 필요합니다: pip install requests")
    sys.exit(1)

# ---------------------------------------------------------------------------
# 선택 의존성: pyautogui (GUI 자동화 — 서버 전용 모드에서는 없어도 됨)
# ---------------------------------------------------------------------------
try:
    import pyautogui
    PYAUTOGUI_AVAILABLE = True
    # 안전장치: 화면 모서리로 마우스 이동 시 자동 중지
    pyautogui.FAILSAFE = True
    pyautogui.PAUSE = 0.3
except ImportError:
    PYAUTOGUI_AVAILABLE = False

# ---------------------------------------------------------------------------
# 선택 의존성: openpyxl (Excel 처리)
# ---------------------------------------------------------------------------
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# ---------------------------------------------------------------------------
# 선택 의존성: pyperclip (클립보드)
# ---------------------------------------------------------------------------
try:
    import pyperclip
    PYPERCLIP_AVAILABLE = True
except ImportError:
    PYPERCLIP_AVAILABLE = False

# ---------------------------------------------------------------------------
# 상수
# ---------------------------------------------------------------------------
DEFAULT_SERVER = "https://sparkling-determination-production-c88b.up.railway.app"
CONFIG_DIR = Path.home() / ".orbit"
CONFIG_FILE = CONFIG_DIR / "config.json"
LOG_FILE = CONFIG_DIR / "auto.log"

VERSION = "1.0.0"

# 주문 텍스트 자동 감지 패턴 (7가지 유형)
ORDER_PATTERNS = [
    # 유형 1: "업체명 / 담당자 : 수량"
    re.compile(r'^.+\s*/\s*.+\s*:\s*\d+', re.MULTILINE),
    # 유형 2: "품목명 x수량" 또는 "품목명 X수량"
    re.compile(r'^.+\s*[xX]\s*\d+', re.MULTILINE),
    # 유형 3: 탭 구분 (업체\t품목\t수량)
    re.compile(r'^.+\t.+\t\d+', re.MULTILINE),
    # 유형 4: "품목 - 수량개" 또는 "품목 - 수량ea"
    re.compile(r'^.+\s*-\s*\d+\s*(개|ea|EA|pcs|PCS)', re.MULTILINE),
    # 유형 5: 콤마 구분 CSV 스타일
    re.compile(r'^.+,.+,\d+', re.MULTILINE),
    # 유형 6: 한글 주문 ("주문", "발주", "오더" 키워드 포함)
    re.compile(r'(주문|발주|오더).+\d+', re.MULTILINE),
    # 유형 7: 괄호 수량 "품목(수량)"
    re.compile(r'^.+\(\d+\)', re.MULTILINE),
]

# Nenova ERP UI 좌표 기본값 (마우스 학습 데이터로 덮어씀)
NENOVA_COORDS = {
    "customer_field": (350, 280),      # 고객명 입력 필드
    "customer_search_btn": (450, 280), # 고객 검색 버튼
    "item_field": (350, 340),          # 품목 검색 필드
    "item_search_btn": (450, 340),     # 품목 검색 버튼
    "quantity_field": (350, 400),      # 수량 입력 필드
    "save_btn": (700, 500),            # 저장 버튼
    "new_order_btn": (100, 60),        # 신규 주문 버튼
    "confirm_popup": (400, 350),       # 확인 팝업 버튼
}


# ===========================================================================
# 설정 관리
# ===========================================================================
class Config:
    """~/.orbit/config.json 설정 파일 관리"""

    def __init__(self):
        self.server = DEFAULT_SERVER
        self.token = ""
        self.verbose = False
        self.nenova_coords = dict(NENOVA_COORDS)
        self._load()

    def _load(self):
        """설정 파일이 있으면 로드"""
        if CONFIG_FILE.exists():
            try:
                with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                    data = json.load(f)
                self.server = data.get("server", self.server)
                self.token = data.get("token", self.token)
                self.verbose = data.get("verbose", self.verbose)
                # 좌표 덮어쓰기
                saved_coords = data.get("nenova_coords", {})
                for key, val in saved_coords.items():
                    if key in self.nenova_coords and isinstance(val, (list, tuple)) and len(val) == 2:
                        self.nenova_coords[key] = tuple(val)
            except (json.JSONDecodeError, IOError) as e:
                print(f"[경고] 설정 파일 로드 실패: {e}")

    def save(self):
        """현재 설정을 파일에 저장"""
        CONFIG_DIR.mkdir(parents=True, exist_ok=True)
        data = {
            "server": self.server,
            "token": self.token,
            "verbose": self.verbose,
            "nenova_coords": {k: list(v) for k, v in self.nenova_coords.items()},
        }
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

    def override(self, args):
        """CLI 인자로 설정 덮어쓰기"""
        if args.server:
            self.server = args.server.rstrip("/")
        if args.token:
            self.token = args.token
        if args.verbose:
            self.verbose = True


# ===========================================================================
# 로깅 설정
# ===========================================================================
def setup_logging(verbose: bool = False):
    """파일 + 콘솔 로깅 설정"""
    CONFIG_DIR.mkdir(parents=True, exist_ok=True)

    log_format = "%(asctime)s [%(levelname)s] %(message)s"
    date_format = "%Y-%m-%d %H:%M:%S"

    # 파일 핸들러 (항상)
    file_handler = logging.FileHandler(str(LOG_FILE), encoding="utf-8")
    file_handler.setLevel(logging.DEBUG)
    file_handler.setFormatter(logging.Formatter(log_format, date_format))

    # 콘솔 핸들러
    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setLevel(logging.DEBUG if verbose else logging.INFO)
    console_handler.setFormatter(logging.Formatter(log_format, date_format))

    logger = logging.getLogger("orbit")
    logger.setLevel(logging.DEBUG)
    logger.addHandler(file_handler)
    logger.addHandler(console_handler)

    return logger


# ===========================================================================
# API 클라이언트
# ===========================================================================
class OrbitAPI:
    """서버 API 통신 클래스"""

    def __init__(self, config: Config, logger: logging.Logger):
        self.config = config
        self.logger = logger
        self.session = requests.Session()
        self.session.headers.update({
            "Content-Type": "application/json",
            "User-Agent": f"OrbitAuto/{VERSION}",
        })
        if config.token:
            self.session.headers["Authorization"] = f"Bearer {config.token}"

    def _url(self, path: str) -> str:
        return f"{self.config.server}{path}"

    def _request(self, method: str, path: str, **kwargs) -> dict:
        """공통 요청 처리 (에러 핸들링 포함)"""
        url = self._url(path)
        self.logger.debug(f"API {method.upper()} {url}")
        try:
            resp = self.session.request(method, url, timeout=30, **kwargs)
            self.logger.debug(f"응답 코드: {resp.status_code}")
            if resp.status_code >= 400:
                self.logger.error(f"API 오류 [{resp.status_code}]: {resp.text[:300]}")
                return {"success": False, "error": f"HTTP {resp.status_code}", "detail": resp.text[:300]}
            # JSON 응답 시도
            try:
                return resp.json()
            except ValueError:
                return {"success": True, "raw": resp.text[:500]}
        except requests.ConnectionError:
            self.logger.error(f"서버 연결 실패: {url}")
            return {"success": False, "error": "연결 실패", "detail": "서버에 연결할 수 없습니다"}
        except requests.Timeout:
            self.logger.error(f"요청 타임아웃: {url}")
            return {"success": False, "error": "타임아웃", "detail": "서버 응답 시간 초과"}
        except requests.RequestException as e:
            self.logger.error(f"요청 예외: {e}")
            return {"success": False, "error": str(e)}

    def get(self, path: str, params=None) -> dict:
        return self._request("get", path, params=params)

    def post(self, path: str, data=None) -> dict:
        return self._request("post", path, json=data)

    # --- 편의 메서드 ---

    def check_status(self) -> dict:
        """서버 상태 확인"""
        os_status = self.get("/api/os/status")
        think_status = self.get("/api/think/status")
        return {"os": os_status, "think": think_status}

    def sync_master_db(self) -> dict:
        """마스터 DB 동기화"""
        return self.post("/api/automation/learn")

    def parse_text(self, text: str) -> dict:
        """텍스트 파싱 요청"""
        return self.post("/api/automation/parse", data={"text": text})

    def get_parsed_orders(self) -> dict:
        """파싱된 주문 목록 조회"""
        return self.get("/api/automation/orders")


# ===========================================================================
# Nenova ERP 자동화
# ===========================================================================
class NenovaAutomation:
    """Nenova ERP UI 자동화 클래스"""

    def __init__(self, api: OrbitAPI, config: Config, logger: logging.Logger, test_mode: bool = True):
        self.api = api
        self.config = config
        self.logger = logger
        self.test_mode = test_mode
        self.coords = config.nenova_coords

    def _click(self, coord_name: str, description: str = ""):
        """좌표 클릭 (test 모드에서는 로그만)"""
        x, y = self.coords[coord_name]
        label = description or coord_name
        if self.test_mode:
            self.logger.info(f"  [테스트] 클릭: {label} ({x}, {y})")
        else:
            if not PYAUTOGUI_AVAILABLE:
                self.logger.error("pyautogui가 설치되지 않아 실제 클릭을 수행할 수 없습니다")
                raise RuntimeError("pyautogui 필요")
            self.logger.info(f"  [실행] 클릭: {label} ({x}, {y})")
            pyautogui.click(x, y)
            time.sleep(0.5)

    def _type_text(self, text: str, description: str = ""):
        """텍스트 입력 (test 모드에서는 로그만)"""
        label = description or "텍스트 입력"
        if self.test_mode:
            self.logger.info(f"  [테스트] 입력: {label} → \"{text}\"")
        else:
            if not PYAUTOGUI_AVAILABLE:
                self.logger.error("pyautogui가 설치되지 않아 실제 입력을 수행할 수 없습니다")
                raise RuntimeError("pyautogui 필요")
            self.logger.info(f"  [실행] 입력: {label} → \"{text}\"")
            pyautogui.typewrite(text, interval=0.05) if text.isascii() else pyautogui.write(text)
            time.sleep(0.3)

    def _press_key(self, key: str, description: str = ""):
        """키 입력"""
        label = description or key
        if self.test_mode:
            self.logger.info(f"  [테스트] 키 입력: {label}")
        else:
            if not PYAUTOGUI_AVAILABLE:
                raise RuntimeError("pyautogui 필요")
            self.logger.info(f"  [실행] 키 입력: {label}")
            pyautogui.press(key)
            time.sleep(0.3)

    def _tab(self, count: int = 1):
        """Tab 키로 다음 필드 이동"""
        for _ in range(count):
            self._press_key("tab", "Tab → 다음 필드")

    def run_order(self):
        """주문 등록 자동화"""
        mode_label = "테스트(드라이런)" if self.test_mode else "실제 실행"
        self.logger.info(f"=== Nenova 주문 등록 시작 [{mode_label}] ===")

        # 1) 서버에서 파싱된 주문 조회
        self.logger.info("서버에서 파싱된 주문 데이터 조회 중...")
        result = self.api.get_parsed_orders()

        if not result.get("success", False) and "error" in result:
            self.logger.error(f"주문 데이터 조회 실패: {result.get('error')}")
            # 데모용 샘플 데이터로 진행
            self.logger.info("데모 데이터로 테스트를 진행합니다...")
            orders = [
                {"customer": "MEL ROSE CHINA", "contact": "Catherine", "item": "Sample Item A", "quantity": 30},
                {"customer": "GOOD TRADING", "contact": "James", "item": "Sample Item B", "quantity": 15},
            ]
        else:
            orders = result.get("orders", result.get("data", []))
            if not orders:
                self.logger.info("처리할 주문이 없습니다")
                return

        self.logger.info(f"총 {len(orders)}건 주문 처리 예정")

        # 2) 각 주문 처리
        for i, order in enumerate(orders, 1):
            customer = order.get("customer", "")
            contact = order.get("contact", "")
            item = order.get("item", "")
            quantity = str(order.get("quantity", 0))

            self.logger.info(f"\n--- 주문 {i}/{len(orders)}: {customer} / {item} x{quantity} ---")

            # 단계 1: 신규 주문 버튼 클릭
            self._click("new_order_btn", "신규 주문 버튼")
            time.sleep(0.3) if not self.test_mode else None

            # 단계 2: 고객명 입력
            self._click("customer_field", "고객명 필드")
            self._type_text(customer, "고객명")

            # 단계 3: 고객 검색
            self._click("customer_search_btn", "고객 검색 버튼")
            time.sleep(1.0) if not self.test_mode else None
            self._press_key("enter", "검색결과 첫 번째 선택")

            # 단계 4: Tab → 품목 필드로 이동
            self._tab(1)

            # 단계 5: 품목 검색
            self._click("item_field", "품목 검색 필드")
            self._type_text(item, "품목명")
            self._click("item_search_btn", "품목 검색 버튼")
            time.sleep(1.0) if not self.test_mode else None
            self._press_key("enter", "검색결과 첫 번째 선택")

            # 단계 6: Tab → 수량 필드로 이동
            self._tab(1)

            # 단계 7: 수량 입력
            self._click("quantity_field", "수량 필드")
            self._type_text(quantity, "수량")

            # 단계 8: 저장
            self._tab(1)
            self._click("save_btn", "저장 버튼")
            time.sleep(0.5) if not self.test_mode else None

            # 단계 9: 확인 팝업 처리
            self._click("confirm_popup", "확인 팝업")

            self.logger.info(f"  → 주문 {i} 처리 완료: {customer} / {item} x{quantity}")

        self.logger.info(f"\n=== 전체 {len(orders)}건 주문 등록 완료 [{mode_label}] ===")

    def run_search(self, keyword: str = ""):
        """품목 검색"""
        self.logger.info(f"=== Nenova 품목 검색 ===")
        if not keyword:
            self.logger.error("검색 키워드를 입력하세요 (--keyword 옵션)")
            return

        self._click("item_field", "품목 검색 필드")
        self._type_text(keyword, f"검색어: {keyword}")
        self._click("item_search_btn", "품목 검색 버튼")
        self.logger.info(f"'{keyword}' 검색 완료")


# ===========================================================================
# Excel 자동화
# ===========================================================================
class ExcelAutomation:
    """Excel 파일 생성/업데이트 자동화"""

    def __init__(self, api: OrbitAPI, logger: logging.Logger):
        self.api = api
        self.logger = logger

    def _ensure_openpyxl(self):
        if not OPENPYXL_AVAILABLE:
            self.logger.error("openpyxl 패키지가 필요합니다: pip install openpyxl")
            raise RuntimeError("openpyxl 필요")

    def _get_orders_data(self) -> list:
        """서버에서 주문 데이터 가져오기"""
        result = self.api.get_parsed_orders()
        if not result.get("success", False) and "error" in result:
            self.logger.warning(f"서버 데이터 조회 실패: {result.get('error')}")
            self.logger.info("샘플 데이터로 진행합니다...")
            return [
                {"customer": "MEL ROSE CHINA", "contact": "Catherine", "item": "Sample A", "quantity": 30, "date": datetime.now().strftime("%Y-%m-%d")},
                {"customer": "GOOD TRADING", "contact": "James", "item": "Sample B", "quantity": 15, "date": datetime.now().strftime("%Y-%m-%d")},
            ]
        return result.get("orders", result.get("data", []))

    def _style_header(self, ws, row: int, columns: int):
        """헤더 행 스타일 적용"""
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        header_font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
        header_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        for col in range(1, columns + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = thin_border

    def _style_data_row(self, ws, row: int, columns: int):
        """데이터 행 스타일 적용"""
        data_font = Font(name="맑은 고딕", size=10)
        data_align = Alignment(vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        for col in range(1, columns + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border

    def generate_deduction(self, filepath: str):
        """차감내역 Excel 생성/업데이트"""
        self._ensure_openpyxl()
        self.logger.info(f"=== 차감내역 Excel 생성: {filepath} ===")

        orders = self._get_orders_data()
        if not orders:
            self.logger.info("처리할 데이터가 없습니다")
            return

        # 기존 파일이 있으면 열기, 없으면 새로 생성
        if os.path.exists(filepath):
            self.logger.info(f"기존 파일 업데이트: {filepath}")
            wb = openpyxl.load_workbook(filepath)
            ws = wb.active
            start_row = ws.max_row + 1
        else:
            self.logger.info(f"새 파일 생성: {filepath}")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "차감내역"

            # 헤더 설정
            headers = ["날짜", "고객사", "담당자", "품목", "수량", "차감구분", "비고"]
            ws.column_dimensions["A"].width = 14
            ws.column_dimensions["B"].width = 25
            ws.column_dimensions["C"].width = 15
            ws.column_dimensions["D"].width = 30
            ws.column_dimensions["E"].width = 10
            ws.column_dimensions["F"].width = 12
            ws.column_dimensions["G"].width = 20

            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            self._style_header(ws, 1, len(headers))
            start_row = 2

        # 데이터 입력
        for i, order in enumerate(orders):
            row = start_row + i
            ws.cell(row=row, column=1, value=order.get("date", datetime.now().strftime("%Y-%m-%d")))
            ws.cell(row=row, column=2, value=order.get("customer", ""))
            ws.cell(row=row, column=3, value=order.get("contact", ""))
            ws.cell(row=row, column=4, value=order.get("item", ""))
            ws.cell(row=row, column=5, value=order.get("quantity", 0))
            ws.cell(row=row, column=6, value="차감")
            ws.cell(row=row, column=7, value="자동생성")
            self._style_data_row(ws, row, 7)

        wb.save(filepath)
        self.logger.info(f"차감내역 저장 완료: {len(orders)}건 → {filepath}")

    def generate_shipping(self, filepath: str):
        """출고내역 Excel 생성"""
        self._ensure_openpyxl()
        self.logger.info(f"=== 출고내역 Excel 생성: {filepath} ===")

        orders = self._get_orders_data()
        if not orders:
            self.logger.info("처리할 데이터가 없습니다")
            return

        # 기존 파일이 있으면 열기, 없으면 새로 생성
        if os.path.exists(filepath):
            self.logger.info(f"기존 파일 업데이트: {filepath}")
            wb = openpyxl.load_workbook(filepath)
            ws = wb.active
            start_row = ws.max_row + 1
        else:
            self.logger.info(f"새 파일 생성: {filepath}")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "출고내역"

            # 헤더 설정
            headers = ["출고일", "고객사", "담당자", "품목", "수량", "출고상태", "송장번호", "비고"]
            ws.column_dimensions["A"].width = 14
            ws.column_dimensions["B"].width = 25
            ws.column_dimensions["C"].width = 15
            ws.column_dimensions["D"].width = 30
            ws.column_dimensions["E"].width = 10
            ws.column_dimensions["F"].width = 12
            ws.column_dimensions["G"].width = 18
            ws.column_dimensions["H"].width = 20

            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            self._style_header(ws, 1, len(headers))
            start_row = 2

        # 데이터 입력
        for i, order in enumerate(orders):
            row = start_row + i
            ws.cell(row=row, column=1, value=order.get("date", datetime.now().strftime("%Y-%m-%d")))
            ws.cell(row=row, column=2, value=order.get("customer", ""))
            ws.cell(row=row, column=3, value=order.get("contact", ""))
            ws.cell(row=row, column=4, value=order.get("item", ""))
            ws.cell(row=row, column=5, value=order.get("quantity", 0))
            ws.cell(row=row, column=6, value="출고대기")
            ws.cell(row=row, column=7, value="")
            ws.cell(row=row, column=8, value="자동생성")
            self._style_data_row(ws, row, 8)

        wb.save(filepath)
        self.logger.info(f"출고내역 저장 완료: {len(orders)}건 → {filepath}")


# ===========================================================================
# 클립보드 감시 모드
# ===========================================================================
class ClipboardWatcher:
    """클립보드 감시 → 자동 파싱"""

    def __init__(self, api: OrbitAPI, logger: logging.Logger):
        self.api = api
        self.logger = logger
        self.last_content = ""

    def _detect_order_type(self, text: str) -> int:
        """주문 텍스트 유형 감지 (0=미감지, 1~7=유형)"""
        for i, pattern in enumerate(ORDER_PATTERNS, 1):
            if pattern.search(text):
                return i
        return 0

    def _display_parsed(self, result: dict):
        """파싱 결과 터미널 출력"""
        print("\n" + "=" * 60)
        print("  파싱 결과")
        print("=" * 60)

        if result.get("success"):
            orders = result.get("orders", result.get("data", []))
            if orders:
                for i, o in enumerate(orders, 1):
                    customer = o.get("customer", "?")
                    item = o.get("item", "?")
                    qty = o.get("quantity", "?")
                    print(f"  [{i}] {customer} → {item} x{qty}")
                print(f"\n  총 {len(orders)}건 감지")
            else:
                print("  주문 데이터 없음")
        else:
            print(f"  파싱 실패: {result.get('error', '알 수 없는 오류')}")

        print("=" * 60 + "\n")

    def watch(self):
        """클립보드 감시 루프 (2초 간격)"""
        if not PYPERCLIP_AVAILABLE:
            self.logger.error("pyperclip 패키지가 필요합니다: pip install pyperclip")
            return

        print("\n" + "=" * 60)
        print("  Orbit AI 클립보드 감시 모드")
        print("  주문 텍스트를 복사하면 자동으로 파싱합니다")
        print("  종료: Ctrl+C")
        print("=" * 60 + "\n")

        self.logger.info("클립보드 감시 시작 (2초 간격)")

        try:
            # 시작 시 현재 클립보드 내용 기록 (중복 방지)
            try:
                self.last_content = pyperclip.paste()
            except Exception:
                self.last_content = ""

            while True:
                time.sleep(2)

                try:
                    current = pyperclip.paste()
                except Exception as e:
                    self.logger.debug(f"클립보드 읽기 실패: {e}")
                    continue

                # 변화 감지
                if current == self.last_content or not current or not current.strip():
                    continue

                self.last_content = current

                # 주문 유형 감지
                order_type = self._detect_order_type(current)
                if order_type > 0:
                    self.logger.info(f"주문 텍스트 감지 (유형 {order_type})")
                    self.logger.debug(f"감지된 텍스트: {current[:100]}...")

                    # 서버로 파싱 요청
                    result = self.api.parse_text(current)
                    self._display_parsed(result)
                else:
                    self.logger.debug("주문 텍스트 아님 — 무시")

        except KeyboardInterrupt:
            print("\n클립보드 감시 종료")
            self.logger.info("클립보드 감시 종료 (사용자 중단)")


# ===========================================================================
# 메인 명령어 핸들러
# ===========================================================================
def cmd_status(api: OrbitAPI, logger: logging.Logger):
    """서버 상태 확인"""
    logger.info("=== 서버 상태 확인 ===")
    result = api.check_status()

    print("\n" + "=" * 60)
    print("  Orbit AI 서버 상태")
    print("=" * 60)

    # OS 상태
    os_data = result.get("os", {})
    if os_data.get("error"):
        print(f"  [OS 엔진] 오류: {os_data['error']}")
    else:
        status = os_data.get("status", os_data.get("raw", "응답 없음"))
        print(f"  [OS 엔진] {status}")

    # Think 상태
    think_data = result.get("think", {})
    if think_data.get("error"):
        print(f"  [사고 엔진] 오류: {think_data['error']}")
    else:
        status = think_data.get("status", think_data.get("raw", "응답 없음"))
        print(f"  [사고 엔진] {status}")

    print(f"\n  서버: {api.config.server}")
    print(f"  시간: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60 + "\n")


def cmd_sync(api: OrbitAPI, logger: logging.Logger):
    """마스터 DB 동기화"""
    logger.info("=== 마스터 DB 동기화 ===")
    result = api.sync_master_db()

    if result.get("success", False) or "error" not in result:
        logger.info("마스터 DB 동기화 완료")
        print("\n  마스터 DB 동기화 완료")
        if result.get("updated"):
            print(f"  업데이트: {result['updated']}건")
        if result.get("synced"):
            print(f"  동기화: {result['synced']}건")
    else:
        logger.error(f"동기화 실패: {result.get('error')}")
        print(f"\n  동기화 실패: {result.get('error')}")


def cmd_parse(api: OrbitAPI, logger: logging.Logger, text: str):
    """텍스트 파싱 테스트"""
    logger.info(f"=== 파싱 테스트 ===")
    logger.info(f"입력: {text}")

    result = api.parse_text(text)

    print("\n" + "=" * 60)
    print("  파싱 결과")
    print("=" * 60)
    print(f"  입력: {text}")
    print("-" * 60)

    if result.get("success", False) or "orders" in result or "data" in result:
        orders = result.get("orders", result.get("data", []))
        if orders:
            for i, o in enumerate(orders, 1):
                customer = o.get("customer", "?")
                contact = o.get("contact", "")
                item = o.get("item", "?")
                qty = o.get("quantity", "?")
                contact_str = f" ({contact})" if contact else ""
                print(f"  [{i}] {customer}{contact_str} → {item} x{qty}")
            print(f"\n  총 {len(orders)}건 파싱 완료")
        else:
            print("  파싱 결과 없음")
            if result.get("raw"):
                print(f"  서버 응답: {result['raw'][:200]}")
    else:
        print(f"  파싱 실패: {result.get('error', '알 수 없는 오류')}")
        if result.get("detail"):
            print(f"  상세: {result['detail'][:200]}")

    print("=" * 60 + "\n")


def cmd_nenova(api: OrbitAPI, config: Config, logger: logging.Logger, action: str, test_mode: bool, keyword: str = ""):
    """Nenova ERP 자동화"""
    if not test_mode and not PYAUTOGUI_AVAILABLE:
        logger.error("실제 실행 모드에는 pyautogui가 필요합니다: pip install pyautogui")
        print("\n  [오류] pyautogui가 설치되지 않았습니다")
        print("  설치: pip install pyautogui")
        print("  또는 --test 모드로 실행하세요\n")
        return

    nenova = NenovaAutomation(api, config, logger, test_mode=test_mode)

    if action == "order":
        nenova.run_order()
    elif action == "search":
        nenova.run_search(keyword)
    else:
        logger.error(f"알 수 없는 Nenova 액션: {action}")
        print(f"\n  [오류] 지원하지 않는 액션: {action}")
        print("  사용 가능: order, search\n")


def cmd_excel(api: OrbitAPI, logger: logging.Logger, action: str, filepath: str):
    """Excel 자동화"""
    if not OPENPYXL_AVAILABLE:
        logger.error("openpyxl 패키지가 필요합니다: pip install openpyxl")
        print("\n  [오류] openpyxl이 설치되지 않았습니다")
        print("  설치: pip install openpyxl\n")
        return

    excel = ExcelAutomation(api, logger)

    if action == "deduction":
        filepath = filepath or "차감내역.xlsx"
        excel.generate_deduction(filepath)
    elif action == "shipping":
        filepath = filepath or "출고내역.xlsx"
        excel.generate_shipping(filepath)
    else:
        logger.error(f"알 수 없는 Excel 액션: {action}")
        print(f"\n  [오류] 지원하지 않는 액션: {action}")
        print("  사용 가능: deduction, shipping\n")


def cmd_watch(api: OrbitAPI, logger: logging.Logger):
    """클립보드 감시 모드"""
    watcher = ClipboardWatcher(api, logger)
    watcher.watch()


# ===========================================================================
# CLI 인자 파서
# ===========================================================================
def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="orbit-auto",
        description="Orbit AI 자동화 CLI 도구 v" + VERSION,
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
사용 예시:
  %(prog)s --target=nenova --action=order --test     드라이런 (입력 시뮬레이션)
  %(prog)s --target=nenova --action=order --run      실제 실행
  %(prog)s --target=excel --action=deduction --file=차감.xlsx
  %(prog)s --target=excel --action=shipping --file=출고.xlsx
  %(prog)s --status                                   서버 상태 확인
  %(prog)s --sync                                     마스터 DB 동기화
  %(prog)s --parse "MEL ROSE CHINA / Catherine : 30"  파서 테스트
  %(prog)s --watch                                    클립보드 감시 모드
        """,
    )

    # 모드 선택 (상호 배타적이지 않으므로 각각 옵션으로)
    parser.add_argument("--target", choices=["nenova", "excel"],
                        help="자동화 대상 (nenova, excel)")
    parser.add_argument("--action", type=str,
                        help="실행 액션 (order, search, deduction, shipping)")
    parser.add_argument("--test", action="store_true",
                        help="테스트 모드 — 드라이런, 실제 동작 안 함")
    parser.add_argument("--run", action="store_true",
                        help="실제 실행 모드")
    parser.add_argument("--file", type=str, default="",
                        help="Excel 파일 경로")
    parser.add_argument("--keyword", type=str, default="",
                        help="검색 키워드 (--action=search 시)")

    # 독립 명령
    parser.add_argument("--status", action="store_true",
                        help="서버 상태 확인")
    parser.add_argument("--sync", action="store_true",
                        help="마스터 DB 동기화")
    parser.add_argument("--parse", type=str, metavar="TEXT",
                        help="텍스트 파싱 테스트")
    parser.add_argument("--watch", action="store_true",
                        help="클립보드 감시 모드")

    # 설정
    parser.add_argument("--server", type=str, default="",
                        help=f"서버 URL (기본값: {DEFAULT_SERVER})")
    parser.add_argument("--token", type=str, default="",
                        help="인증 토큰")
    parser.add_argument("--verbose", "-v", action="store_true",
                        help="상세 로그 출력")

    # 유틸리티
    parser.add_argument("--version", action="version", version=f"%(prog)s {VERSION}")
    parser.add_argument("--save-config", action="store_true",
                        help="현재 설정을 ~/.orbit/config.json에 저장")

    return parser


# ===========================================================================
# 메인
# ===========================================================================
def main():
    parser = build_parser()
    args = parser.parse_args()

    # 인자가 하나도 없으면 도움말 출력
    if len(sys.argv) == 1:
        parser.print_help()
        sys.exit(0)

    # 설정 로드 + 오버라이드
    config = Config()
    config.override(args)

    # 로깅 설정
    logger = setup_logging(verbose=config.verbose or args.verbose)

    # 설정 저장 요청 시
    if args.save_config:
        config.save()
        logger.info(f"설정 저장 완료: {CONFIG_FILE}")
        print(f"\n  설정 저장: {CONFIG_FILE}\n")
        if not any([args.status, args.sync, args.parse, args.watch, args.target]):
            return

    # API 클라이언트 생성
    api = OrbitAPI(config, logger)

    # --- 명령어 라우팅 ---

    if args.status:
        cmd_status(api, logger)
        return

    if args.sync:
        cmd_sync(api, logger)
        return

    if args.parse:
        cmd_parse(api, logger, args.parse)
        return

    if args.watch:
        cmd_watch(api, logger)
        return

    if args.target:
        if not args.action:
            logger.error("--action 옵션이 필요합니다")
            print("\n  [오류] --target 사용 시 --action 을 지정하세요")
            print("  예시: --target=nenova --action=order --test\n")
            sys.exit(1)

        if args.target == "nenova":
            # --test도 --run도 없으면 기본 test 모드
            test_mode = not args.run
            if args.run and args.test:
                logger.warning("--test와 --run이 동시에 지정됨 → 안전을 위해 테스트 모드로 실행")
                test_mode = True
            cmd_nenova(api, config, logger, args.action, test_mode, args.keyword)

        elif args.target == "excel":
            cmd_excel(api, logger, args.action, args.file)

        return

    # 어떤 명령도 매치되지 않은 경우
    parser.print_help()
    sys.exit(0)


if __name__ == "__main__":
    main()
